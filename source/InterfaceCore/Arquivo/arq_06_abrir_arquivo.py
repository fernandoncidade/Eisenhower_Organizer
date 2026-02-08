from PySide6.QtCore import QCoreApplication
from PySide6.QtWidgets import QFileDialog, QMessageBox, QListWidgetItem
from PySide6.QtCore import Qt
import os
from source.utils.LogManager import LogManager
logger = LogManager.get_logger()

def get_text(text):
    return QCoreApplication.translate("InterfaceGrafica", text)

def abrir_arquivo(self):
    path, filt = QFileDialog.getOpenFileName(self.app, self.get_text("Abrir"), os.path.expanduser("~"), "Excel (*.xlsx);;PDF (*.pdf)")
    if not path:
        return

    ext = os.path.splitext(path)[1].lower()

    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook

        except Exception as e:
            logger.error(f"openpyxl não está disponível: {e}", exc_info=True)
            QMessageBox.critical(self.app, self.get_text("Erro"), self.get_text("openpyxl não está disponível."))
            return

        try:
            wb = load_workbook(path, read_only=True)

            def populate_from_sheet(name, lst, completed=False):
                if name in wb.sheetnames:
                    sheet = wb[name]
                    lst.clear()

                    for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
                        val = row[0]

                        if val and str(val).strip():
                            raw_text = str(val).strip()
                            text, date_iso, time_str = self._parse_text_date_time(raw_text)
                            item = QListWidgetItem(raw_text)
                            item.setFlags(item.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                            item.setCheckState(Qt.Checked if completed else Qt.Unchecked)
                            item.setData(Qt.UserRole, {"text": text, "date": date_iso, "time": time_str})
                            tooltip_lines = []

                            if date_iso:
                                try:
                                    from PySide6.QtCore import QDate
                                    qd = QDate.fromString(date_iso, Qt.ISODate)
                                    if qd.isValid():
                                        tooltip_lines.append(f"{self.get_text('Data') or 'Data'}: {qd.toString(self.app.date_input.displayFormat())}")

                                except Exception as e:
                                    logger.error(f"Erro ao definir tooltip de data: {e}", exc_info=True)

                            if time_str:
                                tooltip_lines.append(f"{self.get_text('Horário') or 'Horário'}: {time_str}")

                            if tooltip_lines:
                                item.setToolTip("\n".join(tooltip_lines))

                            try:
                                self.app.insert_task_into_quadrant_list(lst, item)

                            except Exception as e:
                                logger.error(f"Erro ao inserir tarefa na lista do quadrante: {e}", exc_info=True)
                                lst.addItem(item)

            populate_from_sheet("quadrant1", self.app.quadrant1_list, completed=False)
            populate_from_sheet("quadrant1_completed", self.app.quadrant1_completed_list, completed=True)
            populate_from_sheet("quadrant2", self.app.quadrant2_list, completed=False)
            populate_from_sheet("quadrant2_completed", self.app.quadrant2_completed_list, completed=True)
            populate_from_sheet("quadrant3", self.app.quadrant3_list, completed=False)
            populate_from_sheet("quadrant3_completed", self.app.quadrant3_completed_list, completed=True)
            populate_from_sheet("quadrant4", self.app.quadrant4_list, completed=False)
            populate_from_sheet("quadrant4_completed", self.app.quadrant4_completed_list, completed=True)

            self.app.save_tasks()

            try:
                if hasattr(self.app, "calendar_pane") and self.app.calendar_pane:
                    self.app.calendar_pane.calendar_panel.update_task_list()

            except Exception as e:
                logger.error(f"Erro ao atualizar lista de tarefas no calendário: {e}", exc_info=True)

            QMessageBox.information(self.app, self.get_text("Abrir"), self.get_text("Arquivo importado com sucesso."))

        except Exception as e:
            logger.error(f"Erro ao importar arquivo XLSX: {e}", exc_info=True)
            QMessageBox.critical(self.app, self.get_text("Erro"), f"{self.get_text('Erro') or 'Erro'}: {e}")

    elif ext == ".pdf":
        try:
            from PyPDF2 import PdfReader

        except Exception as e:
            logger.error(f"PyPDF2 não está disponível: {e}", exc_info=True)
            QMessageBox.critical(self.app, self.get_text("Erro"), self.get_text("PyPDF2 não está disponível para ler PDF."))
            return

        try:
            reader = PdfReader(path)
            text = ""

            for page in reader.pages:
                try:
                    page_text = page.extract_text() or ""
                    text += page_text + "\n"

                except Exception as e:
                    logger.error(f"Erro ao extrair texto da página do PDF: {e}", exc_info=True)

            import unicodedata, re

            def normalize(s: str) -> str:
                s2 = unicodedata.normalize("NFKD", s)
                s2 = s2.encode("ASCII", "ignore").decode("ASCII")
                s2 = s2.lower()
                s2 = re.sub(r'\s+', ' ', s2).strip()
                return s2

            title_map = {
                normalize("1º Quadrante - Importante e Urgente"): "quadrant1",
                normalize("Concluídas 1º Quadrante"): "quadrant1_completed",
                normalize("2º Quadrante - Importante, mas Não Urgente"): "quadrant2",
                normalize("Concluídas 2º Quadrante"): "quadrant2_completed",
                normalize("3º Quadrante - Não Importante, mas Urgente"): "quadrant3",
                normalize("Concluídas 3º Quadrante"): "quadrant3_completed",
                normalize("4º Quadrante - Não Importante e Não Urgente"): "quadrant4",
                normalize("Concluídas 4º Quadrante"): "quadrant4_completed",
            }

            segments = {k: [] for k in title_map.values()}
            current_key = None
            for raw_line in text.splitlines():
                line = raw_line.strip()

                if not line:
                    continue

                n = normalize(line)

                if n in title_map:
                    current_key = title_map[n]
                    continue

                if current_key:
                    cleaned = re.sub(r'^[\-\u2013\u2014\u2022\u2023\•\*\•\s]+', '', line).strip()
                    if cleaned:
                        segments[current_key].append(cleaned)

            if not any(segments.values()):
                lower = text.lower()
                keys = ["quadrant1", "quadrant1_completed", "quadrant2", "quadrant2_completed",
                        "quadrant3", "quadrant3_completed", "quadrant4", "quadrant4_completed"]

                found_any = False

                for i, key in enumerate(keys):
                    start = lower.find(key + ":")

                    if start != -1:
                        found_any = True
                        end = len(lower)

                        for k in keys[i+1:]:
                            j = lower.find(k + ":", start+1)
                            if j != -1:
                                end = j
                                break

                        segment_text = text[start+len(key)+1:end].strip()

                        for line in segment_text.splitlines():
                            line = line.strip()
                            if line:
                                cleaned = re.sub(r'^[\-\u2013\u2014\u2022\u2023\•\*\s]+', '', line).strip()
                                if cleaned:
                                    segments[key].append(cleaned)

                if not found_any:
                    QMessageBox.warning(self.app, self.get_text("Abrir"), self.get_text("PDF não está no formato compatível."))
                    return

            def populate_from_list(key, lst, completed=False):
                lst.clear()
                items = segments.get(key, [])
                for it in items:
                    raw_text = it.strip()
                    text_only, date_iso, time_str = self._parse_text_date_time(raw_text)
                    item = QListWidgetItem(raw_text)  # mantém exibição
                    item.setFlags(item.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    item.setCheckState(Qt.Checked if completed else Qt.Unchecked)
                    item.setData(Qt.UserRole, {"text": text_only, "date": date_iso, "time": time_str})
                    tooltip_lines = []

                    if date_iso:
                        try:
                            from PySide6.QtCore import QDate
                            qd = QDate.fromString(date_iso, Qt.ISODate)
                            if qd.isValid():
                                tooltip_lines.append(f"{self.get_text('Data') or 'Data'}: {qd.toString(self.app.date_input.displayFormat())}")

                        except Exception as e:
                            logger.error(f"Erro ao definir tooltip de data no PDF: {e}", exc_info=True)

                    if time_str:
                        tooltip_lines.append(f"{self.get_text('Horário') or 'Horário'}: {time_str}")

                    if tooltip_lines:
                        item.setToolTip("\n".join(tooltip_lines))

                    try:
                        self.app.insert_task_into_quadrant_list(lst, item)

                    except Exception as e:
                        logger.error(f"Erro ao inserir tarefa na lista do quadrante (PDF): {e}", exc_info=True)
                        lst.addItem(item)

            populate_from_list("quadrant1", self.app.quadrant1_list, False)
            populate_from_list("quadrant1_completed", self.app.quadrant1_completed_list, True)
            populate_from_list("quadrant2", self.app.quadrant2_list, False)
            populate_from_list("quadrant2_completed", self.app.quadrant2_completed_list, True)
            populate_from_list("quadrant3", self.app.quadrant3_list, False)
            populate_from_list("quadrant3_completed", self.app.quadrant3_completed_list, True)
            populate_from_list("quadrant4", self.app.quadrant4_list, False)
            populate_from_list("quadrant4_completed", self.app.quadrant4_completed_list, True)

            self.app.save_tasks()

            try:
                if hasattr(self.app, "calendar_pane") and self.app.calendar_pane:
                    self.app.calendar_pane.calendar_panel.update_task_list()

            except Exception as e:
                logger.error(f"Erro ao atualizar lista de tarefas no calendário (PDF): {e}", exc_info=True)

            QMessageBox.information(self.app, self.get_text("Abrir"), self.get_text("PDF importado com sucesso."))

        except Exception as e:
            logger.error(f"Erro ao importar arquivo PDF: {e}", exc_info=True)
            QMessageBox.critical(self.app, self.get_text("Erro"), f"{self.get_text('Erro') or 'Erro'}: {e}")

    else:
        QMessageBox.warning(self.app, self.get_text("Abrir"), self.get_text("Formato de arquivo não suportado."))
