from PySide6.QtCore import QCoreApplication
from PySide6.QtWidgets import QFileDialog, QMessageBox
from PySide6.QtCore import Qt
import os
from source.utils.LogManager import LogManager
logger = LogManager.get_logger()

def get_text(text):
    return QCoreApplication.translate("InterfaceGrafica", text)

def salvar_como(self):
    path, filt = QFileDialog.getSaveFileName(self.app, self.get_text("Salvar"), os.path.expanduser("~"), "JSON (*.json);;Excel (*.xlsx);;PDF (*.pdf)")
    if not path:
        return

    def _qenum_to_int(v):
        try:
            return int(v)

        except Exception:
            try:
                return int(getattr(v, "value"))

            except Exception:
                try:
                    return int(v.value())

                except Exception:
                    return None

    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        try:
            from openpyxl import Workbook

        except Exception as e:
            logger.error(f"openpyxl não está disponível para salvar XLSX: {e}", exc_info=True)
            QMessageBox.critical(self.app, self.get_text("Erro"), self.get_text("openpyxl não está disponível para salvar XLSX."))
            return

        try:
            wb = Workbook()

            def write_sheet(name, lst):
                if name in wb.sheetnames:
                    ws = wb[name]

                else:
                    ws = wb.create_sheet(title=name)

                headers = [
                    "text",
                    "date",
                    "time",
                    "file_path",
                    "description",
                    "priority",
                ]
                for col_index, h in enumerate(headers, start=1):
                    ws.cell(row=1, column=col_index, value=h)

                row = 2
                for i in range(lst.count()):
                    item = lst.item(i)
                    if item and (item.flags() & Qt.ItemIsSelectable):
                        data = item.data(Qt.UserRole) or {}
                        text_val = data.get("text", item.text())
                        date_iso = data.get("date")
                        time_str = data.get("time")
                        file_path = data.get("file_path")
                        description = data.get("description")
                        priority = data.get("priority") if data.get("priority") is not None else None

                        values = [
                            text_val,
                            date_iso,
                            time_str,
                            file_path,
                            description,
                            priority,
                        ]

                        for col_index, v in enumerate(values, start=1):
                            ws.cell(row=row, column=col_index, value=v)

                        row += 1

            if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
                ws_default = wb["Sheet"]
                wb.remove(ws_default)

            write_sheet("quadrant1", self.app.quadrant1_list)
            write_sheet("quadrant1_completed", self.app.quadrant1_completed_list)
            write_sheet("quadrant2", self.app.quadrant2_list)
            write_sheet("quadrant2_completed", self.app.quadrant2_completed_list)
            write_sheet("quadrant3", self.app.quadrant3_list)
            write_sheet("quadrant3_completed", self.app.quadrant3_completed_list)
            write_sheet("quadrant4", self.app.quadrant4_list)
            write_sheet("quadrant4_completed", self.app.quadrant4_completed_list)

            wb.save(path)
            QMessageBox.information(self.app, self.get_text("Salvar"), self.get_text("Arquivo salvo com sucesso."))

        except Exception as e:
            logger.error(f"Erro ao salvar arquivo XLSX: {e}", exc_info=True)
            QMessageBox.critical(self.app, self.get_text("Erro"), f"{self.get_text('Erro') or 'Erro'}: {e}")

    elif ext == ".pdf":
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas

        except Exception as e:
            logger.error(f"reportlab não está disponível para salvar PDF: {e}", exc_info=True)
            QMessageBox.critical(self.app, self.get_text("Erro"), self.get_text("reportlab não está disponível para salvar PDF."))
            return

        try:
            c = canvas.Canvas(path, pagesize=A4)
            width, height = A4
            from reportlab.lib.units import cm
            left_margin = 1 * cm
            right_margin = 1 * cm
            top_margin = 1 * cm
            bottom_margin = 1 * cm
            y = height - top_margin
            c.setFont("Helvetica-Bold", 14)
            c.drawString(left_margin, y, "EISENHOWER ORGANIZER")
            y -= 30
            c.setFont("Helvetica-Bold", 12)
            sections = [
                ("1º Quadrante - Importante e Urgente", self.app.quadrant1_list),
                ("Concluídas 1º Quadrante", self.app.quadrant1_completed_list),
                ("2º Quadrante - Importante, mas Não Urgente", self.app.quadrant2_list),
                ("Concluídas 2º Quadrante", self.app.quadrant2_completed_list),
                ("3º Quadrante - Não Importante, mas Urgente", self.app.quadrant3_list),
                ("Concluídas 3º Quadrante", self.app.quadrant3_completed_list),
                ("4º Quadrante - Não Importante e Não Urgente", self.app.quadrant4_list),
                ("Concluídas 4º Quadrante", self.app.quadrant4_completed_list),
            ]
            c.setFont("Helvetica", 10)
            from reportlab.lib.utils import simpleSplit
            from reportlab.pdfbase.pdfmetrics import stringWidth

            def _draw_wrapped(text, indent, current_y, font_name="Helvetica", font_size=10):
                if text is None:
                    return current_y

                c.setFont(font_name, font_size)
                max_width = width - left_margin - right_margin - indent

                if max_width <= 0:
                    max_width = width - left_margin - right_margin

                raw_lines = simpleSplit(text or "", font_name, font_size, max_width)
                lines = []

                for ln in raw_lines:
                    if stringWidth(ln, font_name, font_size) <= max_width:
                        lines.append(ln)

                    else:
                        cur = ""
                        for ch in ln:
                            if stringWidth(cur + ch, font_name, font_size) <= max_width:
                                cur += ch

                            else:
                                if cur:
                                    lines.append(cur)

                                cur = ch

                        if cur:
                            lines.append(cur)

                y_local = current_y

                for ln in lines:
                    if y_local < (bottom_margin + font_size + 4):
                        c.showPage()
                        y_local = height - top_margin
                        c.setFont(font_name, font_size)

                    c.drawString(left_margin + indent, y_local, ln)
                    y_local -= (font_size + 2)

                return y_local

            def _format_item_line(item, data):
                text_val = data.get("text", item.text())
                date_iso = data.get("date")
                time_str = data.get("time")
                date_display = None

                if date_iso:
                    try:
                        from PySide6.QtCore import QDate
                        qd = QDate.fromString(date_iso, Qt.ISODate)
                        if qd.isValid() and hasattr(self.app, "date_input") and self.app.date_input is not None:
                            date_display = qd.toString(self.app.date_input.displayFormat())

                        else:
                            date_display = date_iso

                    except Exception:
                        date_display = date_iso

                parts = []
                parts.append(text_val)
                tail = []

                if date_display:
                    tail.append(date_display)

                if time_str:
                    tail.append(time_str)

                if tail:
                    parts.append(" — " + " ".join(tail))

                return "".join(parts)

            for title, lst in sections:
                if y < (bottom_margin + 40):
                    c.showPage()
                    y = height - top_margin
                    c.setFont("Helvetica", 10)

                c.setFont("Helvetica-Bold", 11)
                c.drawString(left_margin, y, title)
                y -= 16
                c.setFont("Helvetica", 10)
                if lst.count() == 0:
                    c.drawString(left_margin + 10, y, "-")
                    y -= 12

                else:
                    for i in range(lst.count()):
                        item = lst.item(i)
                        if item and (item.flags() & Qt.ItemIsSelectable):
                            data = item.data(Qt.UserRole) or {}

                            main_line = _format_item_line(item, data)
                            y = _draw_wrapped(f"- {main_line}", 10, y, font_name="Helvetica", font_size=10)

                            extra_order = []
                            if data.get("priority") is not None:
                                extra_order.append(("Prioridade", str(data.get("priority"))))

                            if data.get("description"):
                                extra_order.append(("Descrição", data.get("description")))

                            if data.get("file_path"):
                                extra_order.append(("Arquivo", data.get("file_path")))

                            for label, val in extra_order:
                                y = _draw_wrapped(f"{label}: {val}", 20, y, font_name="Helvetica", font_size=9)

                            if y < (bottom_margin + 40):
                                c.showPage()
                                y = height - top_margin
                                c.setFont("Helvetica", 10)

                y -= 8

            c.save()
            QMessageBox.information(self.app, self.get_text("Salvar"), self.get_text("PDF salvo com sucesso."))

        except Exception as e:
            logger.error(f"Erro ao salvar arquivo PDF: {e}", exc_info=True)
            QMessageBox.critical(self.app, self.get_text("Erro"), f"{self.get_text('Erro') or 'Erro'}: {e}")

    elif ext == ".json":
        try:
            import json as _json

            def list_to_full_entries(lst):
                entries = []
                for i in range(lst.count()):
                    item = lst.item(i)
                    if item and (item.flags() & Qt.ItemIsSelectable):
                        data = item.data(Qt.UserRole) or {}
                        entry = {
                            "display_text": item.text(),
                            "text": data.get("text", item.text()),
                            "date": data.get("date"),
                            "time": data.get("time"),
                            "check_state": _qenum_to_int(item.checkState()),
                            "flags": _qenum_to_int(item.flags()),
                            "priority": data.get("priority"),
                            "description": data.get("description"),
                            "file_path": data.get("file_path"),
                            "tooltip": item.toolTip(),
                            "raw_data": data,
                        }
                        entries.append(entry)

                return entries

            full = {
                "meta": {
                    "date_display_format": getattr(self.app, "date_input", None).displayFormat() if hasattr(self.app, "date_input") and self.app.date_input is not None else None,
                    "language": getattr(getattr(self.app, "gerenciador_traducao", None), "obter_idioma_atual", lambda: None)(),
                },
                "quadrant1": list_to_full_entries(self.app.quadrant1_list),
                "quadrant1_completed": list_to_full_entries(self.app.quadrant1_completed_list),
                "quadrant2": list_to_full_entries(self.app.quadrant2_list),
                "quadrant2_completed": list_to_full_entries(self.app.quadrant2_completed_list),
                "quadrant3": list_to_full_entries(self.app.quadrant3_list),
                "quadrant3_completed": list_to_full_entries(self.app.quadrant3_completed_list),
                "quadrant4": list_to_full_entries(self.app.quadrant4_list),
                "quadrant4_completed": list_to_full_entries(self.app.quadrant4_completed_list),
            }

            with open(path, "w", encoding="utf-8") as f:
                _json.dump(full, f, ensure_ascii=False, indent=2)

            QMessageBox.information(self.app, self.get_text("Salvar"), self.get_text("Arquivo salvo com sucesso."))

        except Exception as e:
            logger.error(f"Erro ao salvar arquivo JSON: {e}", exc_info=True)
            QMessageBox.critical(self.app, self.get_text("Erro"), f"{self.get_text('Erro') or 'Erro'}: {e}")

    else:
        QMessageBox.warning(self.app, self.get_text("Salvar"), self.get_text("Extensão não suportada."))
